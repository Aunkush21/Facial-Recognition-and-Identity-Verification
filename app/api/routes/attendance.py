from datetime import date as date_
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import require_role
from app.core.database import get_db
from app.models.attendance import AttendanceRecord, AttendanceStatus
from app.models.identity import Identity
from app.models.user import User, UserRole
from app.schemas.attendance import AttendanceRow, AttendanceSummary

router = APIRouter(prefix="/attendance", tags=["attendance"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_summary(db: Session, date: date_) -> AttendanceSummary:
    stmt = (
        select(
            Identity.id,
            Identity.external_id,
            Identity.full_name,
            AttendanceRecord.status,
            AttendanceRecord.time_recorded,
            AttendanceRecord.confidence,
        )
        .select_from(Identity)
        .outerjoin(
            AttendanceRecord,
            (AttendanceRecord.identity_id == Identity.id) & (AttendanceRecord.date == date),
        )
        .where(Identity.is_active.is_(True))
        .order_by(Identity.full_name)
    )
    rows = db.execute(stmt).all()

    result_rows = []
    present_count = 0
    for identity_id, external_id, full_name, status_, time_recorded, confidence in rows:
        is_present = status_ == AttendanceStatus.PRESENT
        if is_present:
            present_count += 1
        result_rows.append(
            AttendanceRow(
                identity_id=identity_id,
                external_id=external_id,
                full_name=full_name,
                status=AttendanceStatus.PRESENT.value if is_present else AttendanceStatus.ABSENT.value,
                time_recorded=time_recorded,
                confidence=confidence,
            )
        )

    total = len(result_rows)
    return AttendanceSummary(date=date, total=total, present=present_count, absent=total - present_count, rows=result_rows)


@router.get("", response_model=AttendanceSummary)
def get_attendance(
    date: date_ = Query(default_factory=date_.today),
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.ADMIN.value, UserRole.OPERATOR.value)),
) -> AttendanceSummary:
    return _build_summary(db, date)


@router.get("/export")
def export_attendance(
    date: date_ = Query(default_factory=date_.today),
    db: Session = Depends(get_db),
    _: User = Depends(require_role(UserRole.ADMIN.value, UserRole.OPERATOR.value)),
) -> Response:
    """Export the day's attendance as an .xlsx workbook (one row per active identity)."""
    summary = _build_summary(db, date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([f"Attendance for {date.isoformat()}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Present: {summary.present}", f"Absent: {summary.absent}", f"Total: {summary.total}"])
    ws.append([])

    header = ["External ID", "Full Name", "Status", "Time", "Confidence"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in summary.rows:
        ws.append([
            row.external_id,
            row.full_name,
            row.status,
            row.time_recorded.strftime("%H:%M:%S") if row.time_recorded else "",
            round(row.confidence, 4) if row.confidence is not None else "",
        ])

    widths = [18, 28, 12, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"attendance_{date.isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
